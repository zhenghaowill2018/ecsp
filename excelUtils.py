import os

from openpyxl.chart import (BarChart, LineChart, PieChart, Reference,
                            ScatterChart, Series)
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Protection,
                             Side, colors)

#   title格式(title中文,title名称，列宽,列的单元格水平方向对齐,列的单元格垂直方向对齐,自定义格式类型)   例如:('产品编码','asin',20,'left','center'，'yyyy/m/d')
#   values格式为 {'asin':'B00YBP53ZW','stars',4.6}
#   ***注意*** title的名称 与values中的 key 对应
#   ***注意特殊功能*** 可以values中的value传入公式:     例如:'=IF(C{metis_format1}>=4,"1.好评","2.差评")'   metis_format1为标志字符
def compileSheet(sheet,titles,values):
    columnNum=1
    rowNum=1
    columnStr=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','W','U','W','X','Y','Z'
                'AA','AB','AC','AD','AE','AF','AG','AH','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AW','AU','AW','AX','AY','AZ',
                'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BW','BU','BW','BX','BY','BZ']
    #Font(name=u'微软雅黑', bold=True, italic=True, size=24) size
    #bold:是否粗体 italic:是否斜体 size:字体大小 name:字体名称(中文字体加u对其unicode编码) color='FF000000'
    font=Font(name=u'等线')
    #冻结第一行
    sheet.freeze_panes = 'A2'
    for title in titles:
        sheet.column_dimensions[columnStr[columnNum-1]].width = title[2]
        cell=sheet.cell(rowNum,columnNum)
        cell.value=title[0]
        cell.fill=PatternFill("solid", fgColor="4F81BD")
        cell.font=Font(name=u'等线',color='fdfffe')
        cell.border=Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin',color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin',color='000000')
        )
        columnNum=columnNum+1
    rowNum=rowNum+1
    columnNum=1
    for data in values:
        sheet.row_dimensions[rowNum].height = 30
        for title in titles:
            if 'picture'==title[1] and os.path.exists(os.getcwd()+'/image/'+data['asin']+'_1.jpg'):
                try:
                    img = Image(os.getcwd()+'/image/'+data['asin']+'_1.jpg')
                    newsize = (120, 120)
                    img.width, img.height = newsize
                    sheet.add_image(img, 'A'+str(rowNum))
                except Exception as e:
                    asin=data['asin']
                    os.remove(os.getcwd()+'/image/'+data['asin']+'_1.jpg')
                    cell=sheet.cell(rowNum,columnNum)
                    cell.alignment = Alignment(horizontal=title[3], vertical=title[4], wrap_text=True)
                    cell.value='default'
                    print(f'无法解析{asin}图片')
                #单元格设置居中也影响不了图片居中
                #cell=sheet.cell(rowNum,columnNum)
                #cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell=sheet.cell(rowNum,columnNum)
                #if isinstance(title[0],function):
                cell.value=data.get(title[1],None)
                if isinstance(cell.value,str) and title[1]!='content' and title[1]!='split_content':
                    cell.value=cell.value.format(metis_format1=rowNum)
                #cell中数字水平向左，垂直居中
                cell.alignment = Alignment(horizontal=title[3], vertical=title[4], wrap_text=True)
                cell.font=font
                #判断是否是超链接字段
                if 'link_url' in title[1]:
                    cell.font=Font(name=u'等线',u='single',color=colors.BLUE)
                #如果title[4]为None不修改类型，如果title[4]存在，修改单元格类型
                if title[5]:
                    cell.number_format=title[5]
            columnNum=columnNum+1
        columnNum=1
        rowNum=rowNum+1
    #设置过滤器 注意:最后调用才能作用于所有列
    sheet.auto_filter.ref = sheet.dimensions


def scatterCharts(workbook,rows,index_id,title,style):
    """ 
    功能:生成堆叠柱状图
    """ 
    sheet=workbook.create_sheet(index=index_id, title=title)
    for row in rows:
        sheet.append(row)
    chart = BarChart()
    chart.y_axis.title = '评论数量'
    chart.x_axis.title = '日期'
    chart.height = 20 # default is 7.5
    chart.width = 40 # default is 15
    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row, max_col=len(rows[0]))
    cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 5

    chart.type = "col"
    chart.style = style
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    sheet.add_chart(chart, "A1")

def pieChart(workbook,rows,index_id,title):
    """ 
    功能:生成饼状图
    """ 
    sheet=workbook.create_sheet(index=index_id, title=title)
    for row in rows:
        sheet.append(row)
    pie = PieChart()
    pie.height = 20 # default is 7.5
    pie.width = 30 # default is 15
    labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = title
    slice = DataPoint(idx=0, explosion=5)
    pie.series[0].data_points = [slice]
    # Cut the first slice out of the pie
    # slice = DataPoint(idx=0, explosion=20)
    # pie.series[0].data_points = [slice]
    sheet.add_chart(pie, "A1")

def lineChart(workbook, rows, index_id, title,style):
    """ 
    功能:生成折线图
    """ 
    sheet = workbook.create_sheet(index=index_id, title=title)
    for row in rows:
        sheet.append(row)
    chart = LineChart()
    chart.title = title
    chart.style = style
    chart.height = 20 # default is 7.5
    chart.width = 40 # default is 15
    chart.y_axis.title = 'Count'
    chart.y_axis.crossAx = 500
    chart.x_axis = DateAxis(crossAx=100)
    chart.x_axis.number_format = 'eeee-mm'
    chart.x_axis.majorTimeUnit = "years"
    chart.x_axis.title = 'Date'

    # xvalues = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    # for i in range(2, 4):
    #     values = Reference(sheet, min_col=i, min_row=1, max_row=7)
    #     series = Series(values, xvalues, title_from_data=True)
    #     chart.series.append(series)
    # sheet.add_chart(chart, "A1")

    data = Reference(sheet, min_col=2, min_row=1,max_col=len(rows[0]), max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    dates = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.set_categories(dates)

    sheet.add_chart(chart, "A1")
