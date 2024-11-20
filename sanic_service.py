import asyncio
import datetime
import logging
import os
import time

import numpy as np
import openpyxl

import excelUtils
import settings
from db import mysqlUtils

import functions

mysqlDb=mysqlUtils.Mysql(url=settings.MYSQL_URL)
logger = logging.getLogger(f'interface')
local_loop=None

async def normalService(table,fields,condition):
    result=await mysqlDb.find(table,condition,fields)
    return result


async def categoryList(fields,platforms,areas,countrys):
    platforms='","'.join(platforms) if platforms else None
    areas='","'.join(areas) if areas else None
    countrys='","'.join(countrys) if countrys else None

    platform=f'and platform in ("{platforms}") ' if platforms else ''
    area=f'and area in ("{areas}") ' if areas else ''
    country=f'and country in ("{countrys}") ' if countrys else ''

    condition=f'1=1 '+platform+area+country
    result=await mysqlDb.find('channels',condition,fields)
    return result


async def updateNormalService(table,update,condition):
    result=await mysqlDb.update(table,condition,update)
    return result

async def deleteNormalService(table,condition):
    result=await mysqlDb.delete(table,condition)
    return result

async def insertNormalService(table,records):
    result=await mysqlDb.insert(table,records)
    return result

async def upsertNormalService(table,record):
    result=await mysqlDb.upsert(table,record)
    return result

async def upsertNewCategory(table,record):
    result=None
    try:
        await mysqlDb.update("categories_t",{"category_III":record.get("category_III")},{"is_new":1})
        await mysqlDb.upsert(table,record)
        result=1
    except Exception as e:
        await mysqlDb.update("categories_t",{"category_III":record.get("category_III")},{"is_new":0})
        result=0
    return result    

async def salesSumService(start_time,end_time,channels,category_I,category_II,category_III,duty_p,need):
    result={}
    category_I=f'and cp.category_I="{category_I}" ' if category_I else ''
    category_II=f'and cp.category_II="{category_II}" ' if category_II else ''
    category_III=f'and cp.category_III="{category_III}" ' if category_III else ''
    duty_p_sql=''
    duty_p_condition=''
    if duty_p is not None:
        duty_p_sql=",temp_item_duty_p tdp"
        duty_p_condition=f'and tdp.item_no=sf.item_no and tdp.channel_no=sf.channel_no and tdp.duty_p="{duty_p}" '
    channels='","'.join(channels)
    """
        sales_fact sum
    """
    if 'fact' in need:
        fields=['sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
        table='sales_fact sf,products p,categories_t cp'+duty_p_sql
        condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+duty_p_condition
        sf_result=await mysqlDb.find(table,condition,fields)
        result=sf_result[0]
    """
        sales_target sum
    """
    if 'target' in need:
        st_fields=['sum(target_amount) target_amount','sum(target_qty) target_qty']
        st_table='sales_target st,categories_t cp'
        condition=f'st.category_III=cp.category_III and st.ym>="{start_time}" and st.ym<="{end_time}" and st.channel_no in ("{channels}") '+category_I+category_II+category_III
        st_result=await mysqlDb.find(st_table,condition,st_fields)
        result.update(st_result[0])
        if result.get('sales_amount') and result.get('target_amount'):
            result['amount_rate']=round((result.get('sales_amount')/result.get('target_amount'))*100,1)
        if result.get('sales_qty') and result.get('target_qty'):
            result['qty_rate']=round((result.get('sales_qty')/result.get('target_qty'))*100,1)
    return result

"""
    明年小类计划(销售数据)
"""
async def nextYearCategoryTargetSalesService(start_time,end_time,channels,category_I,category_II,category_III):
    category_I=f'and cp.category_I="{category_I}" ' if category_I else ''
    category_II=f'and cp.category_II="{category_II}" ' if category_II else ''
    category_III=f'and cp.category_III="{category_III}" ' if category_III else ''
    channels='","'.join(channels)
    """
        获取sales target明年数据
    """
    st_fields=["DATE_FORMAT(ym, '%m月') ym",'sum(target_amount) target_amount','sum(target_qty) target_qty']
    st_table='sales_target st,categories_t cp'
    st_condition=f'st.category_III=cp.category_III and st.ym>=DATE_ADD("{start_time}",INTERVAL 1 YEAR) and st.ym<=DATE_ADD("{end_time}",INTERVAL 1 YEAR) and st.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    st_result=await mysqlDb.find(st_table,st_condition,st_fields)
    """
        获取sales fact一年数据
    """
    fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    sf_result=await mysqlDb.find(table,condition,fields)
    """
        获取sales fact前一年数据
    """
    fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sales_amount) before_sales_amount','sum(sales_qty) before_sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>=DATE_ADD("{start_time}",INTERVAL -1 YEAR) and sf.ym<=DATE_ADD("{end_time}",INTERVAL -1 YEAR) and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    before_sf_result=await mysqlDb.find(table,condition,fields)
    """
       数据转换格式
    """
    sf_result={x.get('ym'):x for x in sf_result}
    st_result={x.get('ym'):x for x in st_result}
    before_sf_result={x.get('ym'):x for x in before_sf_result}
    months=['01月','02月','03月','04月','05月','06月','07月','08月','09月','10月','11月','12月']

    result={}
    #销售额总计
    sales_amount_total=0
    #去年销售额总计
    before_sales_amount_total=0
    #销量总计
    sales_qty_total=0
    #去年销量总计
    before_sales_qty_total=0
    #目标销售额总计
    target_amount_total=0
    #目标销售额总计
    target_qty_total=0

    for month in months:
        month_result={}
        month_result.update(sf_result.get(month,{}))
        month_result.update(st_result.get(month,{}))
        month_result.update(before_sf_result.get(month,{}))
        month_result['target_qty']=float(month_result.get('target_qty',0.00))
        month_result['sales_qty']=float(month_result.get('sales_qty',0.00))
        month_result['before_sales_qty']=float(month_result.get('before_sales_qty',0.00))
        """
            计算同期成长率
        """
        if month_result.get('sales_qty') and month_result.get('before_sales_qty'):
            diff_value=month_result.get('sales_qty')-month_result.get('before_sales_qty')
            month_result['qty_rate']=round(getDivision(diff_value,month_result.get('before_sales_qty'))*100,1)
        """
            计算销量成长系数
        """
        if month_result.get('target_qty') and month_result.get('sales_qty'):
            month_result['qty_grow_figure']=round(getDivision(month_result.get('target_qty'),month_result.get('sales_qty')),2)
        """
            计算均单价
        """
        if month_result.get('sales_amount') and month_result.get('sales_qty'):
            month_result['averge_price']=round(getDivision(month_result.get('sales_amount'),month_result.get('sales_qty')),1)
        """
            计算目标均单价
        """
        if month_result.get('target_amount') and month_result.get('target_qty'):
            month_result['target_averge_price']=round(getDivision(month_result.get('target_amount'),month_result.get('target_qty'))*100,1)
        """
            计算金额增长系数
        """
        if month_result.get('target_amount') and month_result.get('sales_amount'):
            month_result['amount_grow_figure']=round(getDivision(month_result.get('target_amount'),month_result.get('sales_amount')),2)

        #去年销量总计
        before_sales_qty_total=before_sales_qty_total+float(month_result.get('before_sales_qty',0.00))
        #今年销量总计
        sales_qty_total=sales_qty_total+float(month_result.get('sales_qty',0.00))
        #明年目标销量总计
        target_qty_total=target_qty_total+float(month_result.get('target_qty',0.00))
        #明年目标金额总计
        target_amount_total=target_amount_total+float(month_result.get('target_amount',0.00))
        #今年销售额总计
        sales_amount_total=sales_amount_total+float(month_result.get('sales_amount',0.00))
        result[month]=month_result
    #计算总计
    total_result={}
    #去年销量总计
    total_result['before_sales_qty']=before_sales_qty_total
    #今年销量总计
    total_result['sales_qty']=sales_qty_total
    #总计同期成长率
    total_result['qty_rate']=round(getDivision(sales_qty_total-before_sales_qty_total,before_sales_qty_total)*100,1)
    #目标销量总计
    total_result['target_qty']=target_qty_total
    #总计销量成长系数
    total_result['qty_grow_figure']=round(getDivision(target_qty_total,sales_qty_total),2)
    #今年均单价
    total_result['averge_price']=round(getDivision(sales_amount_total,sales_qty_total),1)
    #目标均单价
    total_result['target_averge_price']=round(getDivision(target_amount_total,target_qty_total),1)
    #目标金额总计
    total_result['target_amount']=target_amount_total
    #今年金额总计
    total_result['sales_amount']=sales_amount_total
    #总计金额成长系数
    total_result['amount_grow_figure']=round(getDivision(target_amount_total,sales_amount_total),2)

    result['total']=total_result
    return result

"""
    明年小类计划
"""
async def nextYearCategoryTargetService(start_time,end_time,channels,category_I,category_II,category_III):
    category_I=f'and cp.category_I="{category_I}" ' if category_I else ''
    category_II=f'and cp.category_II="{category_II}" ' if category_II else ''
    category_III=f'and cp.category_III="{category_III}" ' if category_III else ''
    channels='","'.join(channels)
    """
        获取sales target明年数据
    """
    st_fields=["DATE_FORMAT(ym, '%m月') ym",'sum(target_amount) target_amount','sum(target_qty) target_qty']
    st_table='sales_target st,categories_t cp'
    st_condition=f'st.category_III=cp.category_III and st.ym>=DATE_ADD("{start_time}",INTERVAL 1 YEAR) and st.ym<=DATE_ADD("{end_time}",INTERVAL 1 YEAR) and st.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    st_result=await mysqlDb.find(st_table,st_condition,st_fields)
    """
        获取sales target 今年数据 因后添加,暂取名now_st_result
    """
    st_fields=["DATE_FORMAT(ym, '%m月') ym",'sum(target_amount) now_target_amount','sum(target_qty) now_target_qty']
    st_table='sales_target st,categories_t cp'
    st_condition=f'st.category_III=cp.category_III and st.ym>="{start_time}" and st.ym<="{end_time}" and st.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    now_st_result=await mysqlDb.find(st_table,st_condition,st_fields)
    """
        获取sales fact一年数据
    """
    fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    sf_result=await mysqlDb.find(table,condition,fields)
    """
        获取sales fact前一年数据
    """
    fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sales_amount) before_sales_amount','sum(sales_qty) before_sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>=DATE_ADD("{start_time}",INTERVAL -1 YEAR) and sf.ym<=DATE_ADD("{end_time}",INTERVAL -1 YEAR) and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    before_sf_result=await mysqlDb.find(table,condition,fields)
    """
        获取sales_target_item采购数据
    """
    st_fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sti.target_qty) purchase_target_qty']
    st_table='sales_target_item sti,categories_t cp'
    st_condition=f'sti.category_III=cp.category_III and sti.ym>="{start_time}" and sti.ym<="{end_time}" and sti.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    pc_result=await mysqlDb.find(st_table,st_condition,st_fields)
    """
       数据转换格式
    """
    sf_result={x.get('ym'):x for x in sf_result}
    st_result={x.get('ym'):x for x in st_result}
    now_st_result={x.get('ym'):x for x in now_st_result}
    before_sf_result={x.get('ym'):x for x in before_sf_result}
    pc_result={x.get('ym'):x for x in pc_result}
    months=['01月','02月','03月','04月','05月','06月','07月','08月','09月','10月','11月','12月']

    result={}
    #销售额总计
    sales_amount_total=0
    #去年销售额总计
    before_sales_amount_total=0
    #销量总计
    sales_qty_total=0
    #去年销量总计
    before_sales_qty_total=0
    #目标销售额总计
    target_amount_total=0
    #目标销售额总计
    target_qty_total=0
    #今年目标销量总计
    now_target_amount_total=0
    #今年目标销售额总计
    now_target_qty_total=0
    #采购计划总计
    purchase_result_total=0

    for month in months:
        month_result={}
        month_result.update(sf_result.get(month,{}))
        month_result.update(st_result.get(month,{}))
        month_result.update(before_sf_result.get(month,{}))
        month_result.update(now_st_result.get(month,{}))
        month_result.update(pc_result.get(month,{}))
        month_result['target_qty']=float(month_result.get('target_qty',0.00))
        month_result['sales_qty']=float(month_result.get('sales_qty',0.00))
        month_result['before_sales_qty']=float(month_result.get('before_sales_qty',0.00))
        month_result['now_target_qty']=float(month_result.get('now_target_qty',0.00))
        """
            计算同期成长率
        """
        if month_result.get('sales_qty') and month_result.get('before_sales_qty'):
            diff_value=month_result.get('sales_qty')-month_result.get('before_sales_qty')
            month_result['qty_rate']=round(getDivision(diff_value,month_result.get('before_sales_qty'))*100,1)
        """
            计算销量成长系数
        """
        if month_result.get('target_qty') and month_result.get('sales_qty'):
            month_result['qty_grow_figure']=round(getDivision(month_result.get('target_qty'),month_result.get('sales_qty')),2)
        """
            计算均单价
        """
        if month_result.get('sales_amount') and month_result.get('sales_qty'):
            month_result['averge_price']=round(getDivision(month_result.get('sales_amount'),month_result.get('sales_qty')),1)
        """
            计算目标均单价
        """
        if month_result.get('target_amount') and month_result.get('target_qty'):
            month_result['target_averge_price']=round(getDivision(month_result.get('target_amount'),month_result.get('target_qty'))*100,1)
        """
            计算金额增长系数
        """
        if month_result.get('target_amount') and month_result.get('sales_amount'):
            month_result['amount_grow_figure']=round(getDivision(month_result.get('target_amount'),month_result.get('sales_amount')),2)

        #去年销量总计
        before_sales_qty_total=before_sales_qty_total+float(month_result.get('before_sales_qty',0.00))
        #今年销量总计
        sales_qty_total=sales_qty_total+float(month_result.get('sales_qty',0.00))
        #明年目标销量总计
        target_qty_total=target_qty_total+float(month_result.get('target_qty',0.00))
        #明年目标金额总计
        target_amount_total=target_amount_total+float(month_result.get('target_amount',0.00))
        #今年目标销量总计
        now_target_qty_total=now_target_qty_total+float(month_result.get('now_target_qty',0.00))
        #今年目标金额总计
        now_target_amount_total=now_target_amount_total+float(month_result.get('now_target_amount',0.00))
        #今年销售额总计
        sales_amount_total=sales_amount_total+float(month_result.get('sales_amount',0.00))
        #采购计划总计
        purchase_result_total=purchase_result_total+float(month_result.get('purchase_target_qty',0.00))
        result[month]=month_result
    #计算总计
    total_result={}
    #去年销量总计
    total_result['before_sales_qty']=before_sales_qty_total
    #今年销量总计
    total_result['sales_qty']=sales_qty_total
    #总计同期成长率
    total_result['qty_rate']=round(getDivision(sales_qty_total-before_sales_qty_total,before_sales_qty_total)*100,1)
    #明年目标销量总计
    total_result['target_qty']=target_qty_total
    #今年目标销量总计
    total_result['now_target_qty']=now_target_qty_total
    #总计销量成长系数
    total_result['qty_grow_figure']=round(getDivision(target_qty_total,sales_qty_total),2)
    #今年均单价
    total_result['averge_price']=round(getDivision(sales_amount_total,sales_qty_total),1)
    #去年目标均单价
    total_result['target_averge_price']=round(getDivision(target_amount_total,target_qty_total),1)
    #今年目标均单价
    total_result['now_target_averge_price']=round(getDivision(now_target_amount_total,now_target_qty_total),1)
    #明年目标金额总计
    total_result['target_amount']=target_amount_total
    #今年目标金额总计
    total_result['now_target_amount']=now_target_amount_total
    #今年金额总计
    total_result['sales_amount']=sales_amount_total
    #总计金额成长系数
    total_result['amount_grow_figure']=round(getDivision(target_amount_total,sales_amount_total),2)
    #采购数量总计
    total_result['purchase_target_qty']=purchase_result_total
    result['total']=total_result
    return result

#移动成长率
async def removeGrowRateService(ym,channels,category_III):
    channels='","'.join(channels)
    """
        获取sales fact前三个月销售和
    """
    fields=['sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>=DATE_ADD("{ym}",interval -3 MONTH) and sf.ym<=DATE_ADD("{ym}",interval -1 MONTH) and sf.channel_no in ("{channels}") and p.category_III="{category_III}"'
    sf_result=await mysqlDb.find(table,condition,fields)
    """
        获取sales fact去年前三个月销售和
    """
    fields=['sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>=DATE_ADD("{ym}",interval -15 MONTH) and sf.ym<=DATE_ADD("{ym}",interval -13 MONTH) and sf.channel_no in ("{channels}") and p.category_III="{category_III}"'
    before_sf_result=await mysqlDb.find(table,condition,fields)
    result=None
    if sf_result[0].get('sales_qty') and before_sf_result[0].get('sales_qty'):
        result=round(getDivision(float(sf_result[0].get('sales_qty'))-float(before_sf_result[0].get('sales_qty')),float(before_sf_result[0].get('sales_qty')))*100,1)
    return result

#实际小类占中类比
async def factMidCategoryRateService(start_time,end_time,channels,category_II,category_III):
    channels='","'.join(channels)
    """
        获取sales fact小类的销售总计
    """
    fields=['sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") and cp.category_III="{category_III}" and cp.category_II="{category_II}"'
    category_III_result=await mysqlDb.find(table,condition,fields)
    """
        获取sales fact中类的销售总计
    """
    fields=['sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") and cp.category_II="{category_II}"'
    category_II_result=await mysqlDb.find(table,condition,fields)
    result=None
    if category_III_result[0].get('sales_amount') and category_II_result[0].get('sales_amount'):
        result=round(getDivision(float(category_III_result[0].get('sales_amount',0.00)),float(category_II_result[0].get('sales_amount',0.00)))*100,1)
    return result

#目标小类占中类比
async def targetMidCategoryRateService(start_time,end_time,channels,category_II,category_III):
    channels='","'.join(channels)
    """
        获取sales target小类数据
    """
    st_fields=['sum(target_amount) target_amount','sum(target_qty) target_qty']
    st_table='sales_target st,categories_t cp'
    st_condition=f'st.category_III=cp.category_III and st.ym>="{start_time}" and st.ym<="{end_time}" and st.channel_no in ("{channels}") and cp.category_III="{category_III}" and cp.category_II="{category_II}"'
    category_III_result=await mysqlDb.find(st_table,st_condition,st_fields)

    """
        获取sales target中类数据
    """
    st_fields=['sum(target_amount) target_amount','sum(target_qty) target_qty']
    st_table='sales_target st,categories_t cp'
    st_condition=f'st.category_III=cp.category_III and st.ym>="{start_time}" and st.ym<="{end_time}" and st.channel_no in ("{channels}") and cp.category_II="{category_II}"'
    category_II_result=await mysqlDb.find(st_table,st_condition,st_fields)
    result=None
    if category_III_result[0].get('target_qty') and category_II_result[0].get('target_qty'):
        result=round(getDivision(float(category_III_result[0].get('target_qty',0.00)),float(category_II_result[0].get('target_qty',0.00)))*100,1)
    return result


"""
    明年货号采购计划
"""
async def nextYearItemTargetService(start_time,end_time,channel_no,item_no):
    """
        获取sales fact一年数据
    """
    fields=["DATE_FORMAT(ym, '%m月') ym",'sales_amount','sales_qty']
    table='sales_fact sf'
    condition=f'sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no="{channel_no}" and sf.item_no="{item_no}"'
    sf_result=await mysqlDb.find(table,condition,fields)
    """
        获取sales item target明年数据
    """
    st_fields=["DATE_FORMAT(ym, '%m月') ym",'target_amount','target_qty']
    st_table='sales_target_item sti'
    st_condition=f'sti.ym>=DATE_ADD("{start_time}",INTERVAL 1 YEAR) and sti.ym<=DATE_ADD("{end_time}",INTERVAL -1 YEAR) and sti.channel_no="{channel_no}" and sti.item_no="{item_no}"'
    st_result=await mysqlDb.find(st_table,st_condition,st_fields)
    """
        获取sales fact前一年数据
    """
    before_year_fields=["DATE_FORMAT(ym, '%m月') ym",'sales_amount','sales_qty']
    before_year_condition=f'sf.ym>=DATE_ADD("{start_time}",INTERVAL -1 YEAR) and sf.ym<=DATE_ADD("{end_time}",INTERVAL -1 YEAR) and sf.channel_no="{channel_no}" and sf.item_no="{item_no}"'
    sf_result=await mysqlDb.find(table,condition,fields)
    before_year_result=await mysqlDb.find(table,before_year_condition,before_year_fields)
    """
       数据转换格式
    """
    sf_result={x.get('ym'):x for x in sf_result}
    before_year_result={x.get('ym'):x for x in before_year_result}
    st_result={x.get('ym'):x for x in st_result}
    months=['01月','02月','03月','04月','05月','06月','07月','08月','09月','10月','11月','12月']

    result={}
    #销售额总计
    sales_amount_total=0
    #明年目标销售额总计
    target_amount_total=0
    #去年销售额总计
    before_sales_amount_total=0
    #销量总计
    sales_qty_total=0
    #明年目标销量总计
    target_qty_total=0
    #去年销量总计
    before_sales_qty_total=0
    for month in months:
        month_result={}
        month_result.update(sf_result.get(month,{}))
        month_result.update(before_year_result.get(month,{}))
        month_result.update(st_result.get(month,{}))
        """
            计算目标销售额成长率
        """
        if month_result.get('sales_amount') and month_result.get('target_amount'):
            diff_value=month_result.get('target_amount')-month_result.get('sales_amount')
            month_result['amount_rate']=round(getDivision(diff_value,month_result.get('sales_amount'))*100,1)
        """
            计算目标销量成长率
        """
        if month_result.get('sales_qty') and month_result.get('target_qty'):
            diff_value=month_result.get('target_qty')-month_result.get('sales_qty')
            month_result['qty_rate']=round(getDivision(diff_value,month_result.get('sales_qty'))*100,1)
        """
            计算销售额成长率
        """
        if month_result.get('sales_amount') and month_result.get('before_sales_amount'):
            diff_value=month_result.get('sales_amount')-month_result.get('before_sales_amount')
            month_result['amount_grow_rate']=round(getDivision(diff_value,month_result.get('before_sales_amount'))*100,1)
        """
            计算销量成长率
        """
        if month_result.get('sales_qty') and month_result.get('before_sales_qty'):
            diff_value=int(month_result.get('sales_qty'))-int(month_result.get('before_sales_qty'))
            month_result['qty_grow_rate']=round(getDivision(diff_value,month_result.get('before_sales_qty'))*100,1)
        sales_amount_total=sales_amount_total+month_result.get('sales_amount',0.00)
        target_amount_total=target_amount_total+month_result.get('target_amount',0.00)
        before_sales_amount_total=before_sales_amount_total+month_result.get('before_sales_amount',0.00)
        sales_qty_total=sales_qty_total+month_result.get('sales_qty',0)
        target_qty_total=target_qty_total+month_result.get('target_qty',0)
        before_sales_qty_total=before_sales_qty_total+month_result.get('before_sales_qty',0)
        result[month]=month_result
    #计算总计
    total_result={}
    #明年目标销售额成长率
    next_amount_diff_value=target_amount_total-sales_amount_total
    total_result['next_amount_grow_rate']=round(getDivision(next_amount_diff_value,sales_amount_total)*100,1)
    #计算实际销售额成长率
    amount_diff_value=sales_amount_total-before_sales_amount_total
    total_result['amount_grow_rate']=round(getDivision(amount_diff_value,before_sales_amount_total)*100,1)
    #计算计划销量成长率
    next_qty_diff_value=target_qty_total-sales_qty_total
    total_result['next_qty_grow_rate']=round(getDivision(next_qty_diff_value,sales_qty_total)*100,1)
    #计算实际销量成长率
    qty_diff_value=int(sales_qty_total)-int(before_sales_qty_total)
    total_result['qty_grow_rate']=round(getDivision(qty_diff_value,before_sales_qty_total)*100,1)

    total_result['sales_amount']=round(sales_amount_total,2)
    total_result['target_amount']=target_amount_total
    total_result['before_sales_amount']=before_sales_amount_total
    total_result['sales_qty']=sales_qty_total
    total_result['target_qty']=target_qty_total
    total_result['before_sales_qty']=before_sales_qty_total
    result['total']=total_result
    return result


async def categoryTargetService(start_time,end_time,channels,category_I,category_II,category_III):
    category_I=f'and cp.category_I="{category_I}" ' if category_I else ''
    category_II=f'and cp.category_II="{category_II}" ' if category_II else ''
    category_III=f'and cp.category_III="{category_III}" ' if category_III else ''
    channels='","'.join(channels)
    """
        获取sales fact一年数据
    """
    fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
    table='sales_fact sf,products p,categories_t cp'
    condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    sf_result=await mysqlDb.find(table,condition,fields)
    """
        获取sales fact前一年数据(去年同期)
    """
    before_year_fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sales_amount) before_sales_amount','sum(sales_qty) before_sales_qty']
    before_year_condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>=DATE_ADD("{start_time}",INTERVAL -1 YEAR) and sf.ym<=DATE_ADD("{end_time}",INTERVAL -1 YEAR) and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    before_year_result=await mysqlDb.find(table,before_year_condition,before_year_fields)
    """
        获取sales target一年数据
    """
    st_fields=["DATE_FORMAT(ym, '%m月') ym",'sum(target_amount) target_amount','sum(target_qty) target_qty']
    st_table='sales_target st,categories_t cp'
    st_condition=f'st.category_III=cp.category_III and st.ym>="{start_time}" and st.ym<="{end_time}" and st.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
    st_result=await mysqlDb.find(st_table,st_condition,st_fields)
    """
       数据转换格式
    """
    sf_result={x.get('ym'):x for x in sf_result}
    before_year_result={x.get('ym'):x for x in before_year_result}
    st_result={x.get('ym'):x for x in st_result}
    months=['01月','02月','03月','04月','05月','06月','07月','08月','09月','10月','11月','12月']

    result={}
    #销售额总计
    sales_amount_total=0
    #目标销售额总计
    target_amount_total=0
    #去年目标销售额总计
    before_sales_amount_total=0
    #销量总计
    sales_qty_total=0
    #目标销量总计
    target_qty_total=0
    #去年销量总计
    before_sales_qty_total=0

    for month in months:
        month_result={}
        month_result.update(sf_result.get(month,{}))
        month_result.update(before_year_result.get(month,{}))
        month_result.update(st_result.get(month,{}))
        #计算销售额达成率
        if month_result.get('sales_amount') and month_result.get('target_amount'):
            month_result['amount_rate']=round(getDivision(month_result.get('sales_amount'),month_result.get('target_amount'))*100,1)
        #计算销量达成率
        if month_result.get('sales_qty') and month_result.get('target_qty'):
            month_result['qty_rate']=round(getDivision(month_result.get('sales_qty'),month_result.get('target_qty'))*100,1)
        #计算销售额成长率
        if month_result.get('sales_amount') and month_result.get('before_sales_amount'):
            diff_value=month_result.get('sales_amount')-month_result.get('before_sales_amount')
            month_result['amount_grow_rate']=round(getDivision(diff_value,month_result.get('before_sales_amount'))*100,1)
        #计算销量成长率
        if month_result.get('sales_qty') and month_result.get('before_sales_qty'):
            diff_value=int(month_result.get('sales_qty'))-int(month_result.get('before_sales_qty'))
            month_result['qty_grow_rate']=round(getDivision(diff_value,month_result.get('before_sales_qty'))*100,1)
        #
        if month_result.get('ym'):month_result.pop('ym')
        month_result['fact_price']=round(getDivision(month_result.get('sales_amount'),int(month_result.get('sales_qty',0))),2)
        month_result['target_price']=round(getDivision(month_result.get('target_amount'),int(month_result.get('target_qty',0))),2)

        sales_amount_total=sales_amount_total+month_result.get('sales_amount',0.00)
        target_amount_total=target_amount_total+month_result.get('target_amount',0.00)
        before_sales_amount_total=before_sales_amount_total+month_result.get('before_sales_amount',0.00)
        sales_qty_total=sales_qty_total+month_result.get('sales_qty',0)
        target_qty_total=target_qty_total+month_result.get('target_qty',0)
        before_sales_qty_total=before_sales_qty_total+month_result.get('before_sales_qty',0)
        result[month]=month_result
    #计算总计
    total_result={}
    #计算总计销售额达成率
    total_result['amount_rate']=round(getDivision(sales_amount_total,target_amount_total)*100,1)
    #计算总计销售额成长率
    amount_diff_value=sales_amount_total-before_sales_amount_total
    total_result['amount_grow_rate']=round(getDivision(amount_diff_value,before_sales_amount_total)*100,1)
    #计算总计销量达成率
    total_result['qty_rate']=round(getDivision(sales_qty_total,target_qty_total)*100,1)
    #计算总计销量成长率
    qty_diff_value=int(sales_qty_total)-int(before_sales_qty_total)
    total_result['qty_grow_rate']=round(getDivision(qty_diff_value,before_sales_qty_total)*100,1)

    total_result['sales_amount']=round(sales_amount_total,2)
    total_result['target_amount']=target_amount_total
    total_result['before_sales_amount']=before_sales_amount_total
    total_result['sales_qty']=sales_qty_total
    total_result['target_qty']=target_qty_total
    total_result['before_sales_qty']=before_sales_qty_total
    #计算单价
    total_result['fact_price']=round(getDivision(total_result.get('sales_amount'),int(total_result.get('sales_qty',0))),2)
    total_result['target_price']=round(getDivision(total_result.get('target_amount'),int(total_result.get('target_qty',0))),2)
    result['total']=total_result
    return result

def getDivision(a,b):
    return float(a)/float(b) if a!=0 and b!=0 else 0


async def itemNoTargetService(start_time,end_time,channels,item_no):
    channels='","'.join(channels)
    itemNo=f'and sf.item_no="{item_no}" '
    """
        获取sales fact一年数据
    """
    fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sales_amount) sales_amount','sum(sales_qty) sales_qty']
    table='sales_fact sf'
    condition=f'sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") '+itemNo+'group by ym'
    sf_result=await mysqlDb.find(table,condition,fields)
    """
        获取sales fact前一年数据(去年同期)
    """
    before_year_fields=["DATE_FORMAT(ym, '%m月') ym",'sum(sales_amount) before_sales_amount','sum(sales_qty) before_sales_qty']
    before_year_condition=f'sf.ym>=DATE_ADD("{start_time}",INTERVAL -1 YEAR) and sf.ym<=DATE_ADD("{end_time}",INTERVAL -1 YEAR) and sf.channel_no in ("{channels}") '+itemNo+'group by ym'
    before_year_result=await mysqlDb.find(table,before_year_condition,before_year_fields)
    """
       数据转换格式
    """
    sf_result={x.get('ym'):x for x in sf_result}
    before_year_result={x.get('ym'):x for x in before_year_result}
    months=['01月','02月','03月','04月','05月','06月','07月','08月','09月','10月','11月','12月']
    result={}
    #销售额总计
    sales_amount_total=0
    #目标销售额总计
    target_amount_total=0
    #去年目标销售额总计
    before_sales_amount_total=0
    #销量总计
    sales_qty_total=0
    #目标销量总计
    target_qty_total=0
    #去年销量总计
    before_sales_qty_total=0
    for month in months:
        month_result={}
        month_result.update(sf_result.get(month,{}))
        month_result.update(before_year_result.get(month,{}))
        #计算销售额成长率
        if month_result.get('sales_amount') and month_result.get('before_sales_amount'):
            diff_value=month_result.get('sales_amount')-month_result.get('before_sales_amount')
            month_result['amount_grow_rate']=round(getDivision(diff_value,month_result.get('before_sales_amount'))*100,1)
        #计算销量成长率
        if month_result.get('sales_qty') and month_result.get('before_sales_qty'):
            diff_value=int(month_result.get('sales_qty'))-int(month_result.get('before_sales_qty'))
            month_result['qty_grow_rate']=round(getDivision(diff_value,month_result.get('before_sales_qty'))*100,1)
        if month_result.get('ym'):month_result.pop('ym')
        month_result['fact_price']=round(getDivision(month_result.get('sales_amount'),int(month_result.get('sales_qty',0))),2)


        sales_amount_total=sales_amount_total+month_result.get('sales_amount',0.00)
        before_sales_amount_total=before_sales_amount_total+month_result.get('before_sales_amount',0.00)
        sales_qty_total=sales_qty_total+month_result.get('sales_qty',0)
        before_sales_qty_total=before_sales_qty_total+month_result.get('before_sales_qty',0)
        result[month]=month_result

    #计算总计
    total_result={}
    #计算总计销售额成长率
    amount_diff_value=sales_amount_total-before_sales_amount_total
    total_result['amount_grow_rate']=round(getDivision(amount_diff_value,before_sales_amount_total)*100,1)
    #计算总计销量成长率
    qty_diff_value=int(sales_qty_total)-int(before_sales_qty_total)
    total_result['qty_grow_rate']=round(getDivision(qty_diff_value,before_sales_qty_total)*100,1)

    total_result['sales_amount']=round(sales_amount_total,2)
    total_result['before_sales_amount']=before_sales_amount_total
    total_result['sales_qty']=sales_qty_total
    total_result['before_sales_qty']=before_sales_qty_total
    #计算单价
    total_result['fact_price']=round(getDivision(total_result.get('sales_amount'),int(total_result.get('sales_qty',0))),2)
    result['total']=total_result
    return result

async def exportExcelService(start_time,end_time,channels,fileName,loop):
    workbook=openpyxl.Workbook()
    sheet=workbook.create_sheet("报告",0)
    channels='","'.join(channels)
    global local_loop
    local_loop=loop
    titleList=[('年月','ym',11,'left','center',None),
        ('大类','category_I',7.25,'left','center',None),('中类','category_II',10,'left','center',None),
        ('小类','category_III',10,'left','center',None),('品牌','brand',7,'left','center',None),('区域','area',7,'left','center',None),('国家','country',8,'left','center',None),
        ('平台','platform',10,'left','center',None),('渠道代码','channel_no',10,'right','center',None),('销售金额','sales_amount',10,'right','center',None),
        ('销售数量','sales_qty',7,'right','center','#,##0'),('销售金额(去年)','before_sales_amount',13,'right','center',None),
        ('销售数量(去年)','before_sales_qty',13,'right','center',None),('目标销售金额','target_amount',13,'right','center',None),
        ('目标销售数量','target_qty',13,'right','center',None),('目标采购数量','purchase_target_qty',13,'right','center',None),
        ('目标销售金额(明年)','next_target_amount',15,'right','center',None),('目标销售数量(明年)','next_target_qty',15,'right','center',None),
        ('退货金额','sales_amount_return',10,'right','center',None),('退货数量','sales_qty_return',7,'right','center',None)]
    try:
        #获取数据
        """
            获取sales fact一年数据
        """
        fields=['sf.ym','cp.category_I','cp.category_II','cp.category_III','p.brand','c.area','c.country','c.platform','c.channel_no','sum(sf.sales_amount) sales_amount','sum(sf.sales_qty) sales_qty']
        table='sales_fact sf,products p,categories_t cp,channels c'
        #condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
        condition=f'sf.item_no=p.item_no and p.category_III=cp.category_III and sf.channel_no=c.channel_no and sf.ym<="{end_time}" and sf.ym>="{start_time}" and sf.channel_no in ("{channels}") group by sf.ym,c.channel_no,cp.category_III'
        sf_result=await mysqlDb.find(table,condition,fields)
        """
            获取sales fact_return一年数据
        """
        fields=['sf.ym','cp.category_I','cp.category_II','cp.category_III','p.brand','c.area','c.country','c.platform','c.channel_no','sum(sf.sales_amount) sales_amount_return','sum(sf.sales_qty) sales_qty_return']
        table='sales_fact_return sf,products p,categories_t cp,channels c'
        #condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
        condition=f'sf.item_no=p.item_no and p.category_III=cp.category_III and sf.channel_no=c.channel_no and sf.ym<="{end_time}" and sf.ym>="{start_time}" and sf.channel_no in ("{channels}") group by sf.ym,c.channel_no,cp.category_III'
        sf_return_result=await mysqlDb.find(table,condition,fields)
        """
            获取sales target一年数据
        """
        st_condition=f'select st.ym,cp.category_I,cp.category_II,cp.category_III,cp.brand,c.area,c.country,c.platform,c.channel_no,st.target_amount,st.target_qty from sales_target st,categories_t cp,channels c where st.category_III=cp.category_III and st.channel_no=c.channel_no and st.ym<="{end_time}" and st.ym>="{start_time}"and st.channel_no in ("{channels}")'
        st_result=await mysqlDb.selectDirt(st_condition,args=None)
        """
            获取sales target明年数据
        """
        st_condition=f'select DATE_ADD(st.ym,INTERVAL -1 YEAR) ym,cp.category_I,cp.category_II,cp.category_III,cp.brand,c.area,c.country,c.platform,c.channel_no,st.target_amount next_target_amount,st.target_qty next_target_qty from sales_target st,categories_t cp,channels c where st.category_III=cp.category_III and st.channel_no=c.channel_no and st.ym<=DATE_ADD("{end_time}",INTERVAL 1 YEAR) and st.ym>=DATE_ADD("{start_time}",INTERVAL 1 YEAR) and st.channel_no in ("{channels}")'
        st_next_result=await mysqlDb.selectDirt(st_condition,args=None)
        """
            获取sales_target_item 采购数据
        """
        st_condition=f'select st.ym,cp.category_I,cp.category_II,cp.category_III,cp.brand,c.area,c.country,c.platform,c.channel_no,st.target_qty purchase_target_qty from (select ym,channel_no,category_III,SUM(target_qty) target_qty from sales_target_item GROUP BY channel_no,category_III,ym) st,categories_t cp,channels c where st.category_III=cp.category_III and st.channel_no=c.channel_no and st.ym<="{end_time}" and st.ym>="{start_time}" and st.channel_no in ("{channels}")'
        pt_result=await mysqlDb.selectDirt(st_condition,args=None)
        """
            获取sales fact前一年数据(去年同期)
        """
        before_year_fields=['DATE_ADD(sf.ym,INTERVAL 1 YEAR) ym','cp.category_I','cp.category_II','cp.category_III','p.brand','c.area','c.country','c.platform','c.channel_no','sum(sf.sales_amount) before_sales_amount','sum(sf.sales_qty) before_sales_qty']
        before_table='sales_fact sf,products p,categories_t cp,channels c'
        before_year_condition=f'sf.item_no=p.item_no and p.category_III=cp.category_III and sf.channel_no=c.channel_no and sf.ym<=DATE_ADD("{end_time}",INTERVAL -1 YEAR) and sf.ym>=DATE_ADD("{start_time}",INTERVAL -1 YEAR) and sf.channel_no in ("{channels}") group by sf.ym,c.channel_no,cp.category_III'
        before_year_result=await mysqlDb.find(before_table,before_year_condition,before_year_fields)

        sf_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('category_III'),x),sf_result))
        sf_return_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('category_III'),x),sf_return_result))
        st_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('category_III'),x),st_result))
        st_next_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('category_III'),x),st_next_result))
        pt_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('category_III'),x),pt_result))
        #联立目标金额
        for key,value in st_result.items():
            if key in sf_result:
                sf_result.get(key).update(value)
            else:
                sf_result[key]=value
        #联立明年目标金额
        for key,value in st_next_result.items():
            if key in sf_result:
                sf_result.get(key).update(value)
            else:
                sf_result[key]=value
        #联立退货金额
        for key,value in sf_return_result.items():
            if key in sf_result:
                sf_result.get(key).update(value)
            else:
                sf_result[key]=value
        before_year_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('category_III'),x),before_year_result))
        #联立去年同期
        for key,value in before_year_result.items():
            if key in sf_result:
                sf_result.get(key).update(value)
            else:
                sf_result[key]=value
        #联立采购计划
        for key,value in pt_result.items():
            if key in sf_result:
                sf_result.get(key).update(value)
            else:
                sf_result[key]=value
        result=list(sf_result.values())
        #result=filter(lambda x:(float(getDefualt(x,'sales_amount',0))+float(getDefualt(x,'sales_qty',0))+float(getDefualt(x,'before_sales_amount',0))+float(getDefualt(x,'before_sales_qty',0))+float(getDefualt(x,'target_amount',0))+float(getDefualt(x,'target_qty',0)))!=0,result)
        excelUtils.compileSheet(sheet, titleList,result)
        await local_loop.run_in_executor(None, workbook.save,os.getcwd()+'/target_data/excel/'+fileName+'.xlsx')
    finally:
        workbook.close()

async def excelNormalService(data,sql,titles,sheet_name,fileName,loop):
    workbook=openpyxl.Workbook()
    sheet_name=sheet_name if sheet_name else '报告'
    sheet=workbook.create_sheet(sheet_name,0)
    global local_loop
    local_loop=loop
    if data==None:
        data=await mysqlDb.selectDirt(sql,args={})
    try:
        excelUtils.compileSheet(sheet, titles,data)
        await local_loop.run_in_executor(None, workbook.save,os.getcwd()+'/target_data/excel/'+fileName+'.xlsx')
    finally:
        workbook.close()

async def salesTargetNew(channels,category_III,year):
    st_fields=["DATE_FORMAT(ym, '%m') ym",'target_amount']
    st_table='sales_target st'
    st_condition=f'st.ym>="{year}-01-01" and st.ym<="{year}-12-01" and st.channel_no="{channels}" and category_III="{category_III}"'
    st_result=await mysqlDb.find(st_table,st_condition,st_fields)
    """
       数据转换格式
    """
    st_result={x.get('ym'):x for x in st_result}
    months=['01','02','03','04','05','06','07','08','09','10','11','12']
    end_result={}
    for month in months:
        end_result[month]=0 if st_result.get(month) is None else st_result.get(month).get('target_amount',0) 
    return end_result

async def salesTargetNewCategory(channels,category_III,year):
    st_fields=["DATE_FORMAT(ym, '%m') ym",'target_amount','target_qty','is_new']
    st_table='sales_target st'
    st_condition=f'st.ym>="{year}-01-01" and st.ym<="{year}-12-01" and st.channel_no="{channels}" and category_III="{category_III}"'
    st_result=await mysqlDb.find(st_table,st_condition,st_fields)
    category_II=functions.database.db.read(f"select category_II from categories_t where category_III='{category_III}'").get_one()
    new_result=functions.database.db.read(f"select DATE_FORMAT(st.ym, '%m') ym,st.target_amount from sales_target st where st.ym>='{year}-01-01' and st.ym<='{year}-12-01' and st.channel_no='{channels}' and st.category_III='新品_{category_II.get('category_ii')}'").get_all()
    months=['01','02','03','04','05','06','07','08','09','10','11','12']
    st_result={x.get('ym'):x for x in st_result}
    new_result={x.get('ym'):x for x in new_result}
    end_result={}
    for month in months:
        target_amount=0 if st_result.get(month) is None else st_result.get(month).get('target_amount',0)
        target_qty=0 if st_result.get(month) is None else st_result.get(month).get('target_qty',0)
        surplus_amount=0 if new_result.get(month) is None else new_result.get(month).get('target_amount',0)
        is_new=None if st_result.get(month) is None else st_result.get(month).get('is_new',0)
        month_result={"target_amount":target_amount,"target_qty":target_qty,"surplus_amount":surplus_amount,'is_new':is_new}
        end_result[month]=month_result
    return end_result



async def costRecordsExportExcelService(start_time,end_time,fileName,loop):
    workbook=openpyxl.Workbook()
    sheet=workbook.create_sheet("报告",0)
    global local_loop
    local_loop=loop
    titleList=[('年月','cost_date',11,'left','center',None),
        ('摘要','cost_subject',20,'left','center',None),('消费金额','monetary',15,'left','center',None),('类别','cost_type_p',7,'left','center',None),
        ('子类','cost_type_c',7,'left','center',None),('包含国别','countrys',7,'left','center',None),('渠道','channels',10,'left','center',None)]
    try:
        fields=['cost_date','cost_subject','monetary','cost_type_p','cost_type_c','countrys','channels']
        table='cost_records'
        condition=f'cost_date>="{start_time}" and cost_date<="{end_time}"'
        result=await mysqlDb.find(table,condition,fields)
        for record in result:
           countrys=record.get('countrys')
           record['countrys']=settings.COUNTRYS[countrys.find('1')]
        excelUtils.compileSheet(sheet, titleList,result)
        await local_loop.run_in_executor(None, workbook.save,os.getcwd()+'/target_data/excel/'+fileName+'.xlsx')
    finally:
        workbook.close()



async def salesFactExportExcelService(channels,fileName,loop):
    workbook=openpyxl.Workbook()
    sheet=workbook.create_sheet("报告",0)
    channels='","'.join(channels)
    global local_loop
    local_loop=loop
    titleList=[('年月','ym',11,'left','center',None),('大类','category_I',7.25,'left','center',None),('中类','category_II',10,'left','center',None),
        ('小类','category_III',10,'left','center',None),('货号','item_no',18,'left','center',None),('区域','area',7,'left','center',None),('国家','country',8,'left','center',None),
        ('平台','platform',10,'left','center',None),('渠道代码','channel_no',10,'right','center',None),('销售金额','sales_amount',10,'right','center',None),
        ('销售数量','sales_qty',10,'right','center','#,##0'),('退货金额','sales_amount_return',10,'right','center',None),('退货数量','sales_qty_return',10,'right','center',None)]
    try:
        #获取数据
        """
            获取sales fact数据
        """
        fields=['sf.ym','ct.category_I','ct.category_II','ct.category_III','sf.item_no','sf.sales_qty','sf.sales_amount','c.channel_no','c.platform','c.area','c.country']
        table='sales_fact sf,channels c,products p,categories_t ct'
        condition=f'sf.channel_no in ("{channels}") and sf.channel_no=c.channel_no and p.item_no=sf.item_no and ct.category_III=p.category_III'
        sf_result=await mysqlDb.find(table,condition,fields)
        """
            获取sales fact return数据
        """
        fields=['sf.ym','ct.category_I','ct.category_II','ct.category_III','sf.item_no','sf.sales_qty sales_qty_return','sf.sales_amount sales_amount_return','c.channel_no','c.platform','c.area','c.country']
        table='sales_fact_return sf,channels c,products p,categories_t ct'
        condition=f'sf.channel_no in ("{channels}") and sf.channel_no=c.channel_no and p.item_no=sf.item_no and ct.category_III=p.category_III'
        sf_return_result=await mysqlDb.find(table,condition,fields)
        sf_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('item_no'),x),sf_result))
        sf_return_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('item_no'),x),sf_return_result))
        #联立退货金额
        for key,value in sf_return_result.items():
            if key in sf_result:
                sf_result.get(key).update(value)
            else:
                sf_result[key]=value
        result=list(sf_result.values())
        excelUtils.compileSheet(sheet, titleList,result)
        await local_loop.run_in_executor(None, workbook.save,os.getcwd()+'/target_data/excel/'+fileName+'.xlsx')
    finally:
        workbook.close()

async def productExportExcelService(fileName,loop):
    workbook=openpyxl.Workbook()
    sheet=workbook.create_sheet("报告",0)
    global local_loop
    local_loop=loop
    titleList=[('货号','item_no',20,'left','center',None),('品牌','brand',7.25,'left','center',None),('大类','category_I',10,'left','center',None),
            ('中类','category_II',10,'left','center',None),('小类','category_III',10,'left','center',None),('产品材质','product_material',20,'left','center',None),('产品特点','product_trait',20,'left','center',None),
            ('价格(USD)','price_USD',10,'right','center',None),('价格(RMB)','price_RMB',10,'right','center',None),
            ('产品长度','product_length',8,'right','center',None),('产品宽度','product_wide',8,'right','center',None),('产品高度','product_high',8,'right','center',None),('产品重量','product_weight',8,'right','center',None),('产品其他信息','product_other_message',20,'left','center',None),
            ('产品折叠长度','product_fold_length',13,'right','center',None),('产品折叠宽度','product_fold_wide',13,'right','center',None),('产品折叠高度','product_fold_high',13,'right','center',None),
            ('内部包装长度','internal_packet_length',13,'right','center',None),('内部包装宽度','internal_packet_wide',13,'right','center',None),('内部包装高度','internal_packet_high',13,'right','center',None),
            ('内部包装箱数','internal_box_num',13,'right','center',None),('内部毛重','internal_rough_weight',13,'right','center',None),
            ('外部包装长度','external_packet_length',13,'right','center',None),('外部包装宽度','external_packet_wide',13,'right','center',None),('外部包装高度','external_packet_high',13,'right','center',None),
            ('外部包装箱数','external_packet_num',13,'right','center',None),('外部包装重量','external_rough_weight',13,'right','center',None),('建议宠物信息','suggest_animal_massage',20,'left','center',None),
            ('建议汽车信息','suggest_car_massage',20,'left','center',None),('更新时间','update_time',10,'right','center',None),('采购备注','purchase_description',40,'left','center','#,##0')]
    try:
        #获取数据
        """
            获取sales target
        """
        fields=['ct.*','p.*']
        table='products p,categories_t ct'
        condition=f'p.category_III=ct.category_III'
        result=await mysqlDb.find(table,condition,fields)
        #result=list(result.values())
        excelUtils.compileSheet(sheet, titleList,result)
        await local_loop.run_in_executor(None, workbook.save,os.getcwd()+'/target_data/excel/'+fileName+'.xlsx')
    finally:
        workbook.close()

async def ExportExcelTargetService(fileName,loop):
    print('')  


async def salesTargetItemExportExcelService(channels,start_time,end_time,fileName,loop):
    workbook=openpyxl.Workbook()
    sheet=workbook.create_sheet("报告",0)
    channels='","'.join(channels)
    global local_loop
    local_loop=loop
    titleList=[('年月','ym',11,'left','center',None),('大类','category_I',7.25,'left','center',None),('中类','category_II',10,'left','center',None),
        ('小类','category_III',10,'left','center',None),('货号','item_no',18,'left','center',None),('区域','area',7,'left','center',None),('国家','country',8,'left','center',None),
        ('平台','platform',10,'left','center',None),('渠道代码','channel_no',10,'right','center',None),('当年销量','sales_qty',13,'right','center','#,##0'),('去年销量','before_sales_qty',13,'right','center','#,##0'),
        ('采购计划','target_qty',13,'right','center','#,##0'),('销量预估','shipped_qty',13,'right','center','#,##0'),('采购备注','purchase_description',30,'left','center','#,##0')]
    try:
        #获取数据
        """
            获取sales target
        """
        fields=['sti.ym ym','cp.category_I','cp.category_II','cp.category_III','sti.item_no item_no','p.brand','c.area','c.country','c.platform','c.channel_no','sti.target_qty target_qty','sti.shipped_qty','p.purchase_description']
        table='sales_target_item sti,products p,categories_t cp,channels c'
        condition=f'sti.item_no=p.item_no and sti.ym<=DATE_ADD("{end_time}",INTERVAL 1 YEAR) and sti.ym>=DATE_ADD("{start_time}",INTERVAL 1 YEAR) and p.category_III=cp.category_III and sti.channel_no=c.channel_no and sti.channel_no in ("{channels}")'
        result=await mysqlDb.find(table,condition,fields)

        """
            获取sales fact前一年数据
        """
        fields=['DATE_ADD(sf.ym,INTERVAL 1 YEAR) ym','cp.category_I','cp.category_II','cp.category_III','sf.item_no','p.brand','c.area','c.country','c.platform','c.channel_no','sf.sales_qty before_sales_qty','p.purchase_description']
        table='sales_fact sf,products p,categories_t cp,channels c'
        #condition=f'sf.item_no=p.item_no and cp.category_III=p.category_III and sf.ym>="{start_time}" and sf.ym<="{end_time}" and sf.channel_no in ("{channels}") '+category_I+category_II+category_III+'group by ym'
        condition=f'sf.item_no=p.item_no and p.category_III=cp.category_III and sf.channel_no=c.channel_no and sf.ym<="{end_time}" and sf.ym>="{start_time}" and sf.channel_no in ("{channels}")'
        before_year_result=await mysqlDb.find(table,condition,fields)

        """
            获取sales fact一年数据
        """
        before_year_fields=['sf.ym','cp.category_I','cp.category_II','cp.category_III','sf.item_no','p.brand','c.area','c.country','c.platform','c.channel_no','sf.sales_qty','p.purchase_description']
        before_table='sales_fact sf,products p,categories_t cp,channels c'
        before_year_condition=f'sf.item_no=p.item_no and p.category_III=cp.category_III and sf.channel_no=c.channel_no and sf.ym<=DATE_ADD("{end_time}",INTERVAL 1 YEAR) and sf.ym>=DATE_ADD("{start_time}",INTERVAL 1 YEAR) and sf.channel_no in ("{channels}")'
        sf_result=await mysqlDb.find(before_table,before_year_condition,before_year_fields)

        result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('item_no'),x),result))
        sf_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('item_no'),x),sf_result))
        before_year_result=dict(map(lambda x:(str(x.get('ym'))+"_"+x.get('channel_no')+x.get('item_no'),x),before_year_result))

        #联立当年
        for key,value in sf_result.items():
            if key in result:
                result.get(key).update(value)
            else:
                result[key]=value

        #联立去年同期
        for key,value in before_year_result.items():
            if key in result:
                result.get(key).update(value)
            else:
                result[key]=value

        #result=list(result.values())
        result=list(result.values())
        excelUtils.compileSheet(sheet, titleList,result)
        await local_loop.run_in_executor(None, workbook.save,os.getcwd()+'/target_data/excel/'+fileName+'.xlsx')
    finally:
        workbook.close()


def getDefualt(x,key,defualt):
    return x.get(key,defualt) if x.get(key,defualt) else defualt

async def priceApportionService(start_time,end_time):
    result_list=await priceApportionTask(start_time,end_time)
    try:
        for result in result_list:
            await mysqlDb.upsert('contrast_table',result)
            logger.info(f'关系型数据库插入成功!!!,数据:{result}')
        logger.info(f'关系型数据库,表:{"contrast_table"}插入完成!!!')
        return 'success'
    except Exception as e:
        logger.error(f'遍历dataset失败!!!{e.args}')
        return 'fail'

async def priceApportionTask(start_time,end_time):
    result_list=[]
    database=functions.database
    channels=functions.channels
    result_channels=[c.get('channel_no') for c in channels if c.get('active')==1]
    result_sales=database.db.read(f'select channel_no,sum(sales_amount) money from sales_fact where ym>="{start_time}" and ym<="{end_time}" GROUP BY channel_no').get_all()
    record=database.db.read(f'select * from cost_records where cost_date>="{start_time}" and cost_date<="{end_time}"').get_all()
    #获取所有国家渠道的比例
    # for i,country in enumerate(settings.COUNTRYS):
    #     #该国家的比例
    #     channel_ratios=np.zeros((len(result_channels)))
    #     #获取国家对应的渠道比例
    #     country_channel=list(filter(lambda x:x.get('active')==1 and x.get('country')==country,channels))
    #     country_channel=list(map(lambda x:x.get('channel_no'),country_channel))
    #     country_sales = [x for x in result_sales if x.get('channel_no') in country_channel]
    #     country_total_money=sum(sales.get('money') for sales in country_sales)
    #     #计算对应渠道比例，插入该国家比例列中
    #     for country_sale in country_sales:
    #         country_ratio=country_sale.get('money')/country_total_money
    #         channel_index=result_channels.index(country_sale.get('channel_no'))
    #         channel_ratios[channel_index]=country_ratio
    #     channel_ratio_list[i]=channel_ratios
    #各类别费用分摊到各个国家与渠道
    cnt=np.zeros((len(settings.APPORTION_TYPE),len(settings.COUNTRYS),len(result_channels)))   #(item, country, channel)
    for data in record:
        try:
            channel_ratio_list=np.zeros((len(settings.COUNTRYS),len(result_channels)))
            #获取分摊渠道
            data_channels=data.get('channels').split(',')
            for i,country in enumerate(settings.COUNTRYS):
                #该国家的比例
                channel_ratios=np.zeros((len(result_channels)))
                country_channels=[c.get('channel_no') for c in channels if c.get('active')==1 and country==c.get('country')]
                need_channels=[c for c in data_channels if c in country_channels]
                for country_chan in need_channels:
                    country_ratio=1/len(need_channels)
                    channel_index=result_channels.index(country_chan)
                    channel_ratios[channel_index]=country_ratio
                channel_ratio_list[i]=channel_ratios
            apportion_num=data.get('monetary') #(country)
            flg=np.array([list(map(int,str(data.get('countrys'))))]).T   #(item,country,channel)
            #获取cnt中的索引位置
            index=settings.APPORTION_TYPE.index((data.get('cost_type_p'),data.get('cost_type_c')))
            cty = flg.copy()
            cty[flg==1] = apportion_num / flg.sum()
            ads = channel_ratio_list * cty
            cnt[index, :, :] += ads
        except Exception as e:
            logger.error(f'错误数据:{data},错误原因:{e.args}')
            continue
    #获取不同渠道的总收入
    result_sales_dict={x.get('channel_no'):x for x in result_sales}
    need_apportion_itemNos=database.db.read(f'select item_no,channel_no,sum(sales_amount) total_sales,sum(sales_qty) total_qtys from sales_fact where ym>="{start_time}" and ym<="{end_time}" GROUP BY item_no,channel_no').get_all()
    channel_country={x.get('channel_no'):x.get('country') for x in channels}

    for need_apportion_itemNo in need_apportion_itemNos:
        channel_no=need_apportion_itemNo.get('channel_no')
        item_no=need_apportion_itemNo.get('item_no')
        country=channel_country.get(channel_no)
        country_index=settings.COUNTRYS.index(country)
        channel_index=result_channels.index(need_apportion_itemNo.get('channel_no'))
        channel_total_sales=result_sales_dict.get(channel_no).get('money')
        result_data={'cin':item_no+'_'+channel_no,'item_no':item_no,'channel_no':channel_no}
        for i,apport in enumerate(settings.APPORTION_TYPE):
            #定位np对应的位置,获取金额
            channel_apporitemtion_money=cnt[i][country_index][channel_index]
            #获取对应渠道，货号分摊到的金额
            item_apporitemtion_money=channel_apporitemtion_money*(need_apportion_itemNo.get('total_sales')/channel_total_sales)
            item_apporitemtion_money=item_apporitemtion_money/int(need_apportion_itemNo.get('total_qtys'))
            flied_name=settings.APPORTION_TYPE_NAME.get((apport[0],apport[1]))
            result_data[flied_name]=float(item_apporitemtion_money)
        result_data['appor_version']=end_time.replace('-','')+"-"+start_time.replace('-','')
        result_data['update_time']=time.strftime("%Y-%m-%d", time.localtime())
        result_list.append(result_data)
    return result_list

async def priceTableService(channel,category_III):
    #获取货号列表
    item_no_list=await mysqlDb.find('products',f'category_III="{category_III}"',['item_no','product_cost'])
    item_no_dict={}
    for x in item_no_list:
        item_no_dict[x.get('item_no')]=x.get('product_cost')
    #item_no_list=functions.database.db.read(f'select item_no,purchase_cost from products where category_III="{category_III}"').get_all()
    items_str='","'.join([r.get('item_no') for r in item_no_list])
    #fields=[f"{item_no_list[0].get('purchase_cost')} purchase_cost,item_no,selling_fee,ocean_freight,storage_fee,tariff,vat,express_fee,advertise_fee,profit,returns,min_price,unit_price,manage_other_fee,financial_other_fee"]
    fields=[f"item_no,sale_fee,ocean_freight,tariff,advertise_fee,storage_fee,express_fee,sale_other_fee,manage_other_fee,financial_other_fee,min_price,unit_price"]
    table='price_table'
    condition=f'channel_no="{channel}" and item_no in ("{items_str}")'
    price_result=await mysqlDb.find(table,condition,fields)

    for pr in price_result:
        pr['product_cost']=item_no_dict.get(pr.get('item_no'))
        total_cost=getDefualt(pr,'product_cost',0.00)+getDefualt(pr,'sale_fee',0.00)+getDefualt(pr,'ocean_freight',0.00)+getDefualt(pr,'tariff',0.00)
        +getDefualt(pr,'advertise_fee',0.00)+getDefualt(pr,'storage_fee',0.00)+getDefualt(pr,'express_fee',0.00)+getDefualt(pr,'sale_other_fee',0.00)
        +getDefualt(pr,'manage_other_fee',0.00)
        +getDefualt(pr,'financial_other_fee',0.00)
        if total_cost==0:
            continue
        pr['total_cost']=round(total_cost,2)
        pr['total_rate']=count_rate(pr,total_cost)
        if pr.get('min_price')!=None:
            pr['min_price_rate']=count_rate(pr,getDefualt(pr,'min_price',0))
        if pr.get('unit_price')!=None:
            pr['unit_price_rate']=count_rate(pr,pr.get('unit_price'))
    return price_result


def count_rate(pr,total_cost):
    rate_list={}
    rate_list['product_cost_rate']=int(getDefualt(pr,'product_cost',0.00)*100/total_cost)
    rate_list['sale_fee_rate']=int(getDefualt(pr,'sale_fee',0.00)*100/total_cost)
    rate_list['ocean_freight_rate']=int(getDefualt(pr,'ocean_freight',0.00)*100/total_cost)
    rate_list['tariff_rate']=int(getDefualt(pr,'tariff',0.00)*100/total_cost)
    rate_list['advertise_fee_rate']=int(getDefualt(pr,'advertise_fee',0.00)*100/total_cost)
    rate_list['storage_fee_rate']=int(getDefualt(pr,'storage_fee',0.00)*100/total_cost)
    rate_list['express_fee_rate']=int(getDefualt(pr,'express_fee',0.00)*100/total_cost)
    rate_list['sale_other_fee_rate']=int(getDefualt(pr,'sale_other_fee',0.00)*100/total_cost)
    rate_list['salary_fee_rate']=int(getDefualt(pr,'salary_fee',0.00)*100/total_cost)
    rate_list['office_fee_rate']=int(getDefualt(pr,'office_fee',0.00)*100/total_cost)
    rate_list['rantal_fee_rate']=int(getDefualt(pr,'rantal_fee',0.00)*100/total_cost)
    rate_list['manage_other_fee_rate']=int(getDefualt(pr,'manage_other_fee',0.00)*100/total_cost)
    rate_list['interest_fee_rate']=int(getDefualt(pr,'interest_fee',0.00)*100/total_cost)
    rate_list['financial_other_fee_rate']=int(getDefualt(pr,'financial_other_fee',0.00)*100/total_cost)
    return rate_list

async def priceCompareService(channel,category_III,start_time,end_time,country):
    #获取货号列表
    item_no_list=await mysqlDb.find('products',f'category_III="{category_III}"',['item_no','product_cost'])
    #item_no_list=functions.database.db.read(f'select item_no,purchase_cost from products where category_III="{category_III}"').get_all()
    item_no_dict={}
    for x in item_no_list:
        item_no_dict[x.get('item_no')]=x.get('product_cost')
    items_str='","'.join([r.get('item_no') for r in item_no_list])
    fields=[f"item_no,channel_no,sale_fee,ocean_freight,tariff,VAT,advertise_fee,storage_fee,express_fee,sale_other_fee,manage_other_fee,financial_other_fee"]
    table='price_table'
    condition=f'channel_no="{channel}" and item_no in ("{items_str}")'
    price_result=await mysqlDb.find(table,condition,fields)
    for pr in price_result:
        ct_fields=['*']
        ct_table='contrast_table'
        pr['product_cost']=item_no_dict.get(pr.get('item_no'))
        ct_condition=f'channel_no="{channel}" and item_no in ("{pr.get("item_no")}")'
        ct_result=await mysqlDb.find(ct_table,ct_condition,ct_fields)
        #添加平均佣金
        fields=["sum(mbd.product_sales)/sum(mbd.quantity) average_price","sum(mbd.selling_fees)/sum(mbd.quantity) sale_fee","sum(mbd.product_sales_tax)/sum(mbd.quantity) VAT","sum(mbd.other_transaction_fees)/sum(mbd.quantity) sale_other_fee"]
        table='amazon_business_data mbd,temp_erp_item tei'
        condition=f'mbd.sku=tei.sku and tei.item_no="{pr.get("item_no")}" and type="order" and mbd.country="{country}" and  order_time<="{end_time}" and order_time>="{start_time}"'
        db_result=await mysqlDb.find(table,condition,fields)

        ct=ct_result[0] if len(ct_result)>0 else {}
        if db_result[0].get("sale_fee")!=None:
            ct['sale_fee']=-db_result[0].get("sale_fee") if db_result[0].get("sale_fee")<0 else db_result[0].get("sale_fee")
        if db_result[0].get("VAT")!=None:
            ct['VAT']=-db_result[0].get("VAT") if db_result[0].get("VAT")<0 else db_result[0].get("VAT")
        if db_result[0].get("sale_other_fee")!=None:
            ct['sale_other_fee']=-db_result[0].get("sale_other_fee") if db_result[0].get("sale_other_fee")<0 else db_result[0].get("sale_other_fee")
        pr['contrast_data']=ct
        total_cost=getDefualt(pr,'product_cost',0.00)+getDefualt(pr,'sale_fee',0.00)+getDefualt(pr,'ocean_freight',0.00)+getDefualt(pr,'tariff',0.00)
        +getDefualt(pr,'advertise_fee',0.00)+getDefualt(pr,'storage_fee',0.00)+getDefualt(pr,'express_fee',0.00)+getDefualt(pr,'sale_other_fee',0.00)
        +getDefualt(pr,'manage_other_fee',0.00)
        +getDefualt(pr,'financial_other_fee',0.00)
        pr['total_cost']=round(total_cost,2)
        diff_dirt={}
        diff_dirt['sale_fee']=round(getDefualt(pr,'sale_fee',0.00)-getDefualt(ct,'sale_fee',0.00),2)
        diff_dirt['VAT']=round(getDefualt(pr,'VAT',0.00)-getDefualt(ct,'VAT',0.00),2)
        diff_dirt['ocean_freight']=round(getDefualt(pr,'ocean_freight',0.00)-getDefualt(ct,'ocean_freight',0.00),2)
        diff_dirt['storage_fee']=round(getDefualt(pr,'storage_fee',0.00)-getDefualt(ct,'storage_fee',0.00),2)
        diff_dirt['tariff']=round(getDefualt(pr,'tariff',0.00)-getDefualt(ct,'tariff',0.00),2)
        diff_dirt['advertise_fee']=round(getDefualt(pr,'advertise_fee',0.00)-getDefualt(ct,'advertise_fee',0.00),2)
        #diff_dirt['consumption_tax']=round(getDefualt(pr,'consumption_tax',0.00)-getDefualt(ct,'consumption_tax',0.00),2)
        diff_dirt['storage_fee']=round(getDefualt(pr,'storage_fee',0.00)-getDefualt(ct,'storage_fee',0.00),2)
        diff_dirt['express_fee']=round(getDefualt(pr,'express_fee',0.00)-getDefualt(ct,'express_fee',0.00),2)
        diff_dirt['sale_other_fee']=round(getDefualt(pr,'sale_other_fee',0.00)-getDefualt(ct,'sale_other_fee',0.00),2)
        diff_dirt['manage_other_fee']=round(getDefualt(pr,'manage_other_fee',0.00)-getDefualt(ct,'manage_other_fee',0.00),2)
        diff_dirt['financial_other_fee']=round(getDefualt(pr,'financial_other_fee',0.00)-getDefualt(ct,'financial_other_fee',0.00),2)
        # diff_dirt['manage_other_fee']=round(getDefualt(pr,'manage_other_fee',0.00)-getDefualt(ct,'manage_other_fee',0.00),2)
        # diff_dirt['financial_other_fee']=round(getDefualt(pr,'financial_other_fee',0.00)-getDefualt(ct,'financial_other_fee',0.00),2)
        pr['diff_data']=diff_dirt
    return price_result

async def averagePriceService(item_no,country,start_time,end_time):
    fields=["sum(mbd.product_sales)/sum(mbd.quantity) average_price","sum(mbd.selling_fees)/sum(mbd.quantity) average_sell_fee","sum(mbd.product_sales_tax)/sum(mbd.quantity) average_sales_tax"]
    table='amazon_business_data mbd,temp_erp_item tei'
    condition=f'mbd.sku=tei.sku and tei.item_no="{item_no}" and type="order" and mbd.country="{country}" and  order_time<="{end_time}" and order_time>="{start_time}"'
    price_result=await mysqlDb.find(table,condition,fields)
    return price_result

async def categoryTargetItem(ym,channel,category_III):
    item_no_list=await mysqlDb.find('products',f'category_III="{category_III}"',['item_no'])
    result={}
    for item_no in item_no_list:
        item_no=item_no['item_no']
        before_sales_qty=await mysqlDb.find('sales_fact',f'item_no="{item_no}" and channel_no="{channel}" and ym=DATE_ADD("{ym}",INTERVAL -1 YEAR)',['sales_qty'])
        sales_qty=await mysqlDb.find('sales_fact',f'item_no="{item_no}" and channel_no="{channel}" and ym="{ym}"',['sales_qty'])
        #target_qty=await mysqlDb.find('sales_target_item',f'item_no="{item_no}" and channel_no="{channel}" and ym=DATE_ADD("{ym}",INTERVAL 1 YEAR)',['target_qty','shipped_qty'])
        target_qty=await mysqlDb.find('sales_target_item',f'item_no="{item_no}" and channel_no="{channel}" and ym="{ym}"',['target_qty','shipped_qty'])
        result[item_no]={'ym':ym,'item_no':item_no,'sales_qty':sales_qty[0]['sales_qty'] if len(sales_qty)>0 else 0,
                        'target_qty':target_qty[0]['target_qty'] if len(target_qty)>0 else 0,
                        'shipped_qty':target_qty[0]['shipped_qty'] if len(target_qty)>0 else 0,
                        'before_sales_qty':before_sales_qty[0]['sales_qty'] if len(before_sales_qty)>0 else 0}
    return result

async def temporaryUpdateTarget(start_time,end_time,channel_no):
    database=functions.database
    update_datas=database.db.read(f"select ym,channel_no,category_III,sum(target_qty) target_qty,CURRENT_TIMESTAMP update_time from sales_target_item where ym>='{start_time}' and ym<='{end_time}' and channel_no not in ('500','600','J00','O00','T00') and channel_no='{channel_no}' GROUP BY ym,channel_no,category_III HAVING target_qty>0").get_all()
    update_num=0
    for update_data in update_datas:
        await mysqlDb.upsert('sales_target',update_data)
        update_num=update_num+1
    return {'update_num':update_num}


def getDefualt(x,key,defualt):
    return float(x.get(key,defualt)) if x.get(key,defualt) else defualt

if __name__ == '__main__':
    bbb={'aaa':111,'bbb':None}
    aaa=getDefualt(bbb,'bbb',0)
    print(aaa)
